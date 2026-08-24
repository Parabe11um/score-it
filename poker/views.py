import uuid
from decimal import Decimal, ROUND_HALF_UP

from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Exists, Max, OuterRef, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.cache import never_cache
from django.views.decorators.debug import sensitive_post_parameters
from django.views.decorators.http import require_GET, require_http_methods, require_POST

from .forms import (
    BulkTaskImportForm,
    JoinRoomForm,
    OrganizerRegistrationForm,
    ProjectForm,
    SprintForm,
    VotingSessionForm,
)
from .models import (
    ESTIMATION_GUIDE,
    ESTIMATION_VALUES,
    OrganizerInvitation,
    Participant,
    Project,
    Sprint,
    SprintTask,
    Task,
    Vote,
    VotingRound,
    VotingSession,
    VotingSessionTask,
)


def _format_decimal(value):
    if value is None:
        return None
    rounded = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return format(rounded.normalize(), "f")


def _project_for_user(user, pk):
    return get_object_or_404(Project, pk=pk, owner=user)


def _session_for_user(user, pk):
    return get_object_or_404(
        VotingSession.objects.select_related("project", "current_task"),
        pk=pk,
        project__owner=user,
    )


def _sprint_for_user(user, pk):
    return get_object_or_404(
        Sprint.objects.select_related("project"), pk=pk, project__owner=user
    )


def _task_for_user(user, project_pk, task_pk):
    return get_object_or_404(
        Task,
        pk=task_pk,
        project_id=project_pk,
        project__owner=user,
    )


def _participant_session_key(voting_session):
    return f"score_it_participant_{voting_session.pk}"


def _participant_resume_url(request, voting_session, participant):
    return request.build_absolute_uri(
        reverse(
            "poker:room_resume",
            args=(voting_session.public_token, participant.client_token),
        )
    )


def _participant_rotate_url(voting_session, participant):
    return reverse(
        "poker:participant_resume_rotate",
        args=(voting_session.pk, participant.pk),
    )


def _invitation_unavailable_response(request, invitation):
    if invitation.used_at is not None:
        heading = "Приглашение уже использовано"
        explanation = "По этой ссылке учётная запись организатора уже создана."
    else:
        heading = "Срок приглашения истёк"
        explanation = "Попросите администратора создать новую ссылку."
    return render(
        request,
        "registration/organizer_invitation.html",
        {
            "invitation": invitation,
            "invitation_available": False,
            "invitation_heading": heading,
            "invitation_explanation": explanation,
        },
        status=410,
    )


def _copy_name(name):
    suffix = " — копия"
    return f"{name[:160 - len(suffix)]}{suffix}"


def _warn_if_over_capacity(request, sprint):
    if sprint.is_over_capacity:
        messages.warning(
            request,
            f"Ёмкость спринта превышена на {sprint.capacity_overage_display} points.",
        )


def _participant_from_request(request, voting_session):
    token = request.session.get(_participant_session_key(voting_session))
    if not token:
        return None
    return Participant.objects.filter(
        session=voting_session, client_token=token
    ).first()


def _add_tasks_to_queue(voting_session, tasks):
    existing_ids = set(
        voting_session.queue_items.values_list("task_id", flat=True)
    )
    position = (
        voting_session.queue_items.aggregate(value=Max("position"))["value"] or 0
    )
    queue_items = []
    for task in tasks:
        if task.pk in existing_ids:
            continue
        position += 1
        queue_items.append(
            VotingSessionTask(
                session=voting_session,
                task=task,
                position=position,
            )
        )
    VotingSessionTask.objects.bulk_create(queue_items)
    return queue_items


def _reset_participant_completions(voting_session):
    return voting_session.participants.filter(
        completed_at__isnull=False
    ).update(completed_at=None)


def _create_round_for_item(voting_session, queue_item):
    last_number = (
        VotingRound.objects.filter(
            session=voting_session, task=queue_item.task
        ).aggregate(value=Max("number"))["value"]
        or 0
    )
    voting_round = VotingRound.objects.create(
        session=voting_session,
        task=queue_item.task,
        number=last_number + 1,
    )
    queue_item.status = VotingSessionTask.Status.ACTIVE
    queue_item.completed_at = None
    queue_item.current_round = voting_round
    queue_item.save(
        update_fields=("status", "completed_at", "current_round")
    )
    _reset_participant_completions(voting_session)
    return voting_round


def _open_pending_queue_items(voting_session):
    queue_items = list(
        voting_session.queue_items.select_related("task").filter(
            status__in=(
                VotingSessionTask.Status.PENDING,
                VotingSessionTask.Status.SKIPPED,
            )
        )
    )
    for queue_item in queue_items:
        _create_round_for_item(voting_session, queue_item)

    first_item = voting_session.queue_items.select_related("task").filter(
        status=VotingSessionTask.Status.ACTIVE
    ).first()
    if first_item:
        current_is_open = voting_session.queue_items.filter(
            task_id=voting_session.current_task_id,
            status=VotingSessionTask.Status.ACTIVE,
        ).exists()
        if not current_is_open:
            voting_session.current_task = first_item.task
        voting_session.status = VotingSession.Status.ACTIVE
        voting_session.save(update_fields=("current_task", "status"))
    return queue_items


def _focus_next_open_item(voting_session, current_position):
    next_item = (
        voting_session.queue_items.select_related("task")
        .filter(
            position__gt=current_position,
            status=VotingSessionTask.Status.ACTIVE,
        )
        .first()
    )
    if next_item is None:
        next_item = (
            voting_session.queue_items.select_related("task")
            .filter(status=VotingSessionTask.Status.ACTIVE)
            .first()
        )
    voting_session.current_task = next_item.task if next_item else None
    voting_session.save(update_fields=("current_task",))
    return next_item


def _queue_context(voting_session):
    queue_items = list(
        voting_session.queue_items.select_related("task", "current_round").annotate(
            vote_count=Count("current_round__votes")
        )
    )
    completed = sum(
        item.status == VotingSessionTask.Status.COMPLETED for item in queue_items
    )
    pending = sum(
        item.status == VotingSessionTask.Status.PENDING for item in queue_items
    )
    active = sum(
        item.status == VotingSessionTask.Status.ACTIVE for item in queue_items
    )
    skipped = sum(
        item.status == VotingSessionTask.Status.SKIPPED for item in queue_items
    )
    current_position = next(
        (
            index
            for index, item in enumerate(queue_items, start=1)
            if item.task_id == voting_session.current_task_id
        ),
        None,
    )
    for index, item in enumerate(queue_items, start=1):
        item.display_position = index
        item.is_current = item.task_id == voting_session.current_task_id
    focusable_items = [
        item for item in queue_items if item.status == VotingSessionTask.Status.ACTIVE
    ]
    focus_index = next(
        (
            index
            for index, item in enumerate(focusable_items)
            if item.task_id == voting_session.current_task_id
        ),
        None,
    )
    return {
        "items": queue_items,
        "total": len(queue_items),
        "completed": completed,
        "pending": pending,
        "active": active,
        "skipped": skipped,
        "current_position": current_position,
        "has_previous": focus_index is not None and focus_index > 0,
        "has_next": focus_index is not None
        and focus_index < len(focusable_items) - 1,
    }


def _queue_summary(voting_session):
    summary = voting_session.queue_items.aggregate(
        total=Count("id"),
        completed=Count(
            "id", filter=Q(status=VotingSessionTask.Status.COMPLETED)
        ),
        pending=Count("id", filter=Q(status=VotingSessionTask.Status.PENDING)),
        active=Count("id", filter=Q(status=VotingSessionTask.Status.ACTIVE)),
        skipped=Count("id", filter=Q(status=VotingSessionTask.Status.SKIPPED)),
    )
    summary["current_position"] = (
        voting_session.queue_items.filter(task_id=voting_session.current_task_id)
        .values_list("position", flat=True)
        .first()
    )
    return summary


def _participant_queue_item(voting_session, participant):
    queue_items = voting_session.queue_items.select_related("task", "current_round")
    queue_item = None
    if participant.current_task_id:
        queue_item = queue_items.filter(task_id=participant.current_task_id).first()
    if queue_item is None:
        queue_item = queue_items.filter(
            status=VotingSessionTask.Status.ACTIVE
        ).first() or queue_items.first()
        if queue_item:
            participant.current_task = queue_item.task
            participant.save(update_fields=("current_task",))
    return queue_item


def _participant_queue_summary(voting_session, participant, queue_item):
    queue_rows = list(
        voting_session.queue_items.annotate(
            participant_vote_count=Count(
                "current_round__votes",
                filter=Q(current_round__votes__participant=participant),
                distinct=True,
            )
        ).values(
            "id",
            "task_id",
            "position",
            "status",
            "participant_vote_count",
        ).order_by("position")
    )
    current_index = next(
        (
            index
            for index, row in enumerate(queue_rows)
            if queue_item and row["id"] == queue_item.pk
        ),
        None,
    )
    missing_rows = [
        row for row in queue_rows if row["participant_vote_count"] == 0
    ]
    voted = len(queue_rows) - len(missing_rows)
    total = len(queue_rows)
    current_row = queue_rows[current_index] if current_index is not None else None
    first_missing = missing_rows[0] if missing_rows else None
    return {
        "total": total,
        "completed": sum(
            row["status"] == VotingSessionTask.Status.COMPLETED
            for row in queue_rows
        ),
        "voted": voted,
        "missing": total - voted,
        "all_voted": bool(total and voted == total),
        "current_position": current_row["position"] if current_row else None,
        "current_has_voted": bool(
            current_row and current_row["participant_vote_count"]
        ),
        "has_previous": current_index is not None and current_index > 0,
        "has_next": current_index is not None
        and current_index < total - 1,
        "first_missing_task_id": (
            first_missing["task_id"] if first_missing else None
        ),
        "first_missing_position": (
            first_missing["position"] if first_missing else None
        ),
    }


def _participant_progress(voting_session, current_round_voted_ids=None):
    queue_total = voting_session.queue_items.count()
    current_round_voted_ids = current_round_voted_ids or set()
    participants = list(
        voting_session.participants.annotate(
            progress_voted=Count(
                "votes__voting_round__queue_item",
                filter=Q(
                    votes__voting_round__queue_item__session=voting_session
                ),
                distinct=True,
            )
        )
    )
    for participant in participants:
        participant.progress_total = queue_total
        participant.current_round_voted = participant.pk in current_round_voted_ids
        if participant.progress_voted == 0:
            participant.progress_status = "not_started"
            participant.progress_label = f"0 из {queue_total} · не приступил"
        elif participant.progress_voted >= queue_total and queue_total:
            if participant.completed_at:
                participant.progress_status = "completed"
                participant.progress_label = (
                    f"{queue_total} из {queue_total} · завершил"
                )
            else:
                participant.progress_status = "all_voted"
                participant.progress_label = (
                    f"{queue_total} из {queue_total} · готов"
                )
        else:
            participant.progress_status = "in_progress"
            participant.progress_label = (
                f"{participant.progress_voted} из {queue_total}"
            )
    return participants


def _project_detail_context(
    project,
    *,
    task_filter="all",
    competency_filter="all",
    show_archived=False,
    task_import_form=None,
    session_form=None,
    sprint_form=None,
):
    task_filter_labels = (
        ("all", "Все"),
        ("new", "Новые"),
        ("unestimated", "Не оценены"),
        ("estimated", "Оценены"),
        ("sprint", "В спринте"),
        ("completed", "Завершены"),
    )
    valid_filters = {value for value, _label in task_filter_labels}
    if task_filter not in valid_filters:
        task_filter = "all"

    competency_filter_labels = (
        ("all", "Все направления"),
        ("analysis", "Аналитика"),
        ("development", "Разработка"),
        ("testing", "Тестирование"),
        ("untyped", "Без типа"),
    )
    valid_competency_filters = {
        value for value, _label in competency_filter_labels
    }
    if competency_filter not in valid_competency_filters:
        competency_filter = "all"

    tasks = project.tasks.annotate(
        in_sprint=Exists(
            SprintTask.objects.filter(
                task_id=OuterRef("pk"), status=SprintTask.Status.PLANNED
            )
        ),
        has_voting_history=Exists(
            VotingSessionTask.objects.filter(task_id=OuterRef("pk"))
        ),
    )
    task_queries = {
        "all": Q(),
        "new": Q(
            completed_at__isnull=True,
            status=Task.Status.UNESTIMATED,
            has_voting_history=False,
        ),
        "unestimated": Q(
            completed_at__isnull=True,
            status=Task.Status.UNESTIMATED,
        ),
        "estimated": Q(
            completed_at__isnull=True,
            status=Task.Status.ESTIMATED,
        ),
        "sprint": Q(completed_at__isnull=True, in_sprint=True),
        "completed": Q(completed_at__isnull=False),
    }
    filter_counts = {
        value: tasks.filter(task_queries[value]).count()
        for value, _label in task_filter_labels
    }
    tasks_by_status = tasks.filter(task_queries[task_filter])
    competency_queries = {
        "all": Q(),
        "analysis": Q(competency=Task.Competency.ANALYSIS),
        "development": Q(competency=Task.Competency.DEVELOPMENT),
        "testing": Q(competency=Task.Competency.TESTING),
        "untyped": Q(competency=Task.Competency.NONE),
    }
    competency_filter_counts = {
        value: tasks_by_status.filter(competency_queries[value]).count()
        for value, _label in competency_filter_labels
    }
    filtered_tasks = tasks_by_status.filter(
        competency_queries[competency_filter]
    )

    sessions = project.voting_sessions.select_related("current_task").filter(
        archived_at__isnull=not show_archived
    )
    sprints = project.sprints.filter(
        archived_at__isnull=not show_archived
    )
    return {
        "project": project,
        "tasks": filtered_tasks,
        "task_total": project.tasks.count(),
        "task_filter": task_filter,
        "task_filters": [
            {"value": value, "label": label, "count": filter_counts[value]}
            for value, label in task_filter_labels
        ],
        "competency_filter": competency_filter,
        "competency_filters": [
            {
                "value": value,
                "label": label,
                "count": competency_filter_counts[value],
            }
            for value, label in competency_filter_labels
        ],
        "competency_choices": Task.Competency.choices,
        "sessions": sessions,
        "session_total": project.voting_sessions.count(),
        "sprints": sprints,
        "sprint_total": project.sprints.count(),
        "show_archived": show_archived,
        "archived_total": (
            project.voting_sessions.filter(archived_at__isnull=False).count()
            + project.sprints.filter(archived_at__isnull=False).count()
        ),
        "task_import_form": task_import_form or BulkTaskImportForm(),
        "session_form": session_form or VotingSessionForm(project=project),
        "sprint_form": sprint_form or SprintForm(),
    }


def _dashboard_context(user, project_form=None):
    projects = list(
        Project.objects.filter(owner=user)
        .prefetch_related("tasks", "voting_sessions", "sprints")
        .order_by("-updated_at")
    )
    active_rooms = list(
        VotingSession.objects.filter(
            project__owner=user,
            status=VotingSession.Status.ACTIVE,
            archived_at__isnull=True,
        )
        .select_related("project")
        .annotate(
            participant_count=Count("participants", distinct=True),
            completed_participant_count=Count(
                "participants",
                filter=Q(participants__completed_at__isnull=False),
                distinct=True,
            ),
            queue_total=Count("queue_items", distinct=True),
            queue_completed=Count(
                "queue_items",
                filter=Q(queue_items__status=VotingSessionTask.Status.COMPLETED),
                distinct=True,
            ),
        )
        .order_by("-created_at")
    )
    for voting_session in active_rooms:
        voting_session.incomplete_participant_count = (
            voting_session.participant_count
            - voting_session.completed_participant_count
        )

    active_sprints = list(
        Sprint.objects.filter(
            project__owner=user,
            status=Sprint.Status.ACTIVE,
            archived_at__isnull=True,
        )
        .select_related("project")
        .order_by("end_date", "-created_at")
    )
    upcoming_sprints = list(
        Sprint.objects.filter(
            project__owner=user,
            status__in=(Sprint.Status.PLANNING, Sprint.Status.ACTIVE),
            archived_at__isnull=True,
            end_date__isnull=False,
        )
        .select_related("project")
        .order_by("end_date")[:5]
    )
    unestimated_tasks = Task.objects.filter(
        project__owner=user,
        completed_at__isnull=True,
        status=Task.Status.UNESTIMATED,
    )
    new_task_count = (
        unestimated_tasks.annotate(
            has_voting_history=Exists(
                VotingSessionTask.objects.filter(task_id=OuterRef("pk"))
            )
        )
        .filter(has_voting_history=False)
        .count()
    )
    return {
        "projects": projects,
        "project_form": project_form or ProjectForm(),
        "active_rooms": active_rooms,
        "active_sprints": active_sprints,
        "upcoming_sprints": upcoming_sprints,
        "incomplete_participant_count": sum(
            item.incomplete_participant_count for item in active_rooms
        ),
        "new_task_count": new_task_count,
        "unestimated_task_count": unestimated_tasks.count(),
        "today": timezone.localdate(),
    }


@sensitive_post_parameters("password1", "password2")
@never_cache
@require_http_methods(["GET", "POST"])
def organizer_invitation_accept(request, token):
    invitation = get_object_or_404(OrganizerInvitation, token=token)
    if not invitation.is_available:
        return _invitation_unavailable_response(request, invitation)

    form = OrganizerRegistrationForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        with transaction.atomic():
            invitation = OrganizerInvitation.objects.select_for_update().get(
                pk=invitation.pk
            )
            if not invitation.is_available:
                return _invitation_unavailable_response(request, invitation)

            organizer = form.save()
            invitation.used_at = timezone.now()
            invitation.used_by = organizer
            invitation.save(update_fields=("used_at", "used_by"))

        login(
            request,
            organizer,
            backend="django.contrib.auth.backends.ModelBackend",
        )
        messages.success(
            request,
            "Учётная запись организатора создана. Можно создавать первый проект.",
        )
        return redirect("poker:dashboard")

    return render(
        request,
        "registration/organizer_invitation.html",
        {
            "invitation": invitation,
            "invitation_available": True,
            "registration_form": form,
        },
    )


@login_required
def dashboard(request):
    return render(
        request,
        "poker/dashboard.html",
        _dashboard_context(request.user),
    )


@login_required
@require_POST
def project_create(request):
    form = ProjectForm(request.POST)
    if form.is_valid():
        project = form.save(commit=False)
        project.owner = request.user
        project.save()
        messages.success(request, f"Проект «{project.name}» создан.")
        return redirect(project)

    return render(
        request,
        "poker/dashboard.html",
        _dashboard_context(request.user, project_form=form),
        status=400,
    )


@login_required
def project_detail(request, pk):
    project = _project_for_user(request.user, pk)
    return render(
        request,
        "poker/project_detail.html",
        _project_detail_context(
            project,
            task_filter=request.GET.get("tasks", "all"),
            competency_filter=request.GET.get("competency", "all"),
            show_archived=request.GET.get("archive") == "1",
        ),
    )


@login_required
@require_POST
def task_import(request, pk):
    project = _project_for_user(request.user, pk)
    form = BulkTaskImportForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Проверьте формат списка задач.")
        return render(
            request,
            "poker/project_detail.html",
            _project_detail_context(project, task_import_form=form),
            status=400,
        )

    created_count = 0
    updated_count = 0
    with transaction.atomic():
        for number, title, competency in form.parsed_tasks:
            task, created = Task.objects.get_or_create(
                project=project,
                number=number,
                defaults={"title": title, "competency": competency},
            )
            if created:
                created_count += 1
            elif task.title != title or task.competency != competency:
                task.title = title
                task.competency = competency
                task.save(
                    update_fields=("title", "competency", "updated_at")
                )
                updated_count += 1

    messages.success(
        request,
        f"Импорт завершён: добавлено {created_count}, обновлено {updated_count}.",
    )
    return redirect(project)


@login_required
@require_POST
def task_complete(request, pk, task_pk):
    project = _project_for_user(request.user, pk)
    task = _task_for_user(request.user, project.pk, task_pk)
    if task.completed_at is None:
        task.completed_at = timezone.now()
        task.save(update_fields=("completed_at", "updated_at"))
        messages.success(request, f"Задача {task.number} отмечена завершённой.")
    return redirect(project)


@login_required
@require_POST
def task_reopen(request, pk, task_pk):
    project = _project_for_user(request.user, pk)
    task = _task_for_user(request.user, project.pk, task_pk)
    if task.completed_at is not None:
        task.completed_at = None
        task.save(update_fields=("completed_at", "updated_at"))
        messages.info(request, f"Задача {task.number} возвращена в работу.")
    return redirect(project)


@login_required
@require_POST
def task_competency_update(request, pk, task_pk):
    project = _project_for_user(request.user, pk)
    task = _task_for_user(request.user, project.pk, task_pk)
    redirect_url = project.get_absolute_url()
    if request.GET:
        redirect_url = f"{redirect_url}?{request.GET.urlencode()}"
    competency = request.POST.get("competency", "")
    if competency not in Task.Competency.values:
        messages.error(request, "Неизвестный тип задачи.")
        return redirect(redirect_url)

    if task.competency != competency:
        task.competency = competency
        task.save(update_fields=("competency", "updated_at"))
        messages.success(
            request,
            f"Для задачи {task.number} установлен тип «{task.get_competency_display()}».",
        )
    return redirect(redirect_url)


@login_required
@require_POST
def task_delete(request, pk, task_pk):
    project = _project_for_user(request.user, pk)
    task = _task_for_user(request.user, project.pk, task_pk)
    task_number = task.number
    task.delete()
    messages.success(request, f"Задача {task_number} удалена.")
    return redirect(project)


@login_required
@require_POST
def session_create(request, pk):
    project = _project_for_user(request.user, pk)
    form = VotingSessionForm(request.POST, project=project)
    if form.is_valid():
        with transaction.atomic():
            voting_session = form.save(commit=False)
            voting_session.project = project
            voting_session.save()
            queued_items = _add_tasks_to_queue(
                voting_session,
                project.tasks.filter(pk__in=form.cleaned_data["task_ids"]),
            )
        messages.success(
            request,
            f"Комната создана. В очередь добавлено задач: {len(queued_items)}.",
        )
        return redirect(voting_session)
    messages.error(request, "Не удалось создать комнату: проверьте параметры.")
    return render(
        request,
        "poker/project_detail.html",
        _project_detail_context(project, session_form=form),
        status=400,
    )


@login_required
def session_manage(request, pk):
    voting_session = _session_for_user(request.user, pk)
    current_round = voting_session.current_round
    queue = _queue_context(voting_session)
    queued_task_ids = [item.task_id for item in queue["items"]]
    available_tasks = voting_session.project.tasks.filter(
        completed_at__isnull=True
    ).exclude(pk__in=queued_task_ids)
    public_url = request.build_absolute_uri(voting_session.get_public_url())
    summary = current_round.summary() if current_round else None
    current_round_voted_ids = set(
        current_round.votes.values_list("participant_id", flat=True)
    ) if current_round else set()
    participant_progress = _participant_progress(
        voting_session, current_round_voted_ids
    )
    for participant in participant_progress:
        participant.resume_url = _participant_resume_url(
            request, voting_session, participant
        )
        participant.rotate_url = _participant_rotate_url(
            voting_session, participant
        )
    return render(
        request,
        "poker/session_manage.html",
        {
            "voting_session": voting_session,
            "queue": queue,
            "available_tasks": available_tasks,
            "current_round": current_round,
            "summary": summary,
            "participant_progress": participant_progress,
            "completed_participant_count": sum(
                item.progress_status == "completed"
                for item in participant_progress
            ),
            "average_display": _format_decimal(summary["average"])
            if summary
            else None,
            "public_url": public_url,
        },
    )


@login_required
@require_POST
def session_queue_add(request, pk):
    voting_session = _session_for_user(request.user, pk)
    if voting_session.status == VotingSession.Status.FINISHED:
        messages.error(request, "В завершённую сессию нельзя добавлять задачи.")
        return redirect(voting_session)

    task_ids = request.POST.getlist("task_ids")
    tasks = voting_session.project.tasks.filter(
        pk__in=task_ids, completed_at__isnull=True
    )
    with transaction.atomic():
        created_items = _add_tasks_to_queue(voting_session, tasks)
        if voting_session.status == VotingSession.Status.ACTIVE:
            _open_pending_queue_items(voting_session)
    created = len(created_items)
    if created:
        messages.success(request, f"В очередь добавлено задач: {created}.")
    else:
        messages.info(request, "Выберите хотя бы одну новую задачу.")
    return redirect(voting_session)


@login_required
@require_POST
def session_start(request, pk):
    voting_session = _session_for_user(request.user, pk)
    if voting_session.status == VotingSession.Status.FINISHED:
        messages.error(request, "Завершённую сессию нельзя продолжить.")
        return redirect(voting_session)
    if voting_session.status == VotingSession.Status.ACTIVE:
        messages.info(request, "Голосование уже открыто для всех задач очереди.")
        return redirect(voting_session)

    with transaction.atomic():
        opened_items = _open_pending_queue_items(voting_session)
    if not opened_items and not voting_session.queue_items.filter(
        status=VotingSessionTask.Status.ACTIVE
    ).exists():
        messages.info(request, "В очереди нет задач, ожидающих оценки.")
    else:
        messages.success(
            request,
            "Асинхронное голосование открыто для всех задач очереди.",
        )
    return redirect(voting_session)


@login_required
@require_POST
def session_navigate(request, pk, direction):
    voting_session = _session_for_user(request.user, pk)
    if direction not in ("previous", "next"):
        messages.error(request, "Неизвестное направление перехода.")
        return redirect(voting_session)

    queue_items = list(
        voting_session.queue_items.select_related("task").filter(
            status=VotingSessionTask.Status.ACTIVE
        )
    )
    current_index = next(
        (
            index
            for index, item in enumerate(queue_items)
            if item.task_id == voting_session.current_task_id
        ),
        None,
    )
    if current_index is None:
        target = queue_items[0] if queue_items else None
    else:
        target_index = current_index + (1 if direction == "next" else -1)
        target = (
            queue_items[target_index]
            if 0 <= target_index < len(queue_items)
            else None
        )
    if target is None:
        messages.info(request, "В этом направлении больше нет открытых задач.")
        return redirect(voting_session)

    voting_session.current_task = target.task
    voting_session.save(update_fields=("current_task",))
    return redirect(voting_session)


@login_required
@require_POST
def session_start_task(request, pk, task_pk):
    voting_session = _session_for_user(request.user, pk)
    task = get_object_or_404(Task, pk=task_pk, project=voting_session.project)

    if voting_session.status == VotingSession.Status.FINISHED:
        messages.error(request, "Завершённую сессию нельзя продолжить.")
        return redirect(voting_session)
    with transaction.atomic():
        queue_item = voting_session.queue_items.filter(task=task).first()
        if queue_item is None:
            _add_tasks_to_queue(voting_session, [task])
            queue_item = voting_session.queue_items.get(task=task)
        if queue_item.current_round_id is None or queue_item.current_round.status in (
            VotingRound.Status.CLOSED,
            VotingRound.Status.CANCELLED,
        ):
            _create_round_for_item(voting_session, queue_item)
        voting_session.current_task = task
        voting_session.status = VotingSession.Status.ACTIVE
        voting_session.save(update_fields=("current_task", "status"))

    messages.success(request, f"Открыто голосование по задаче {task.number}.")
    return redirect(voting_session)


@login_required
@require_POST
def session_reveal(request, pk):
    voting_session = _session_for_user(request.user, pk)
    voting_round = voting_session.current_round
    if not voting_round or voting_round.status != VotingRound.Status.VOTING:
        messages.error(request, "Нет активного голосования для раскрытия.")
        return redirect(voting_session)
    vote_count = voting_round.votes.count()
    if vote_count < voting_session.minimum_participants:
        remaining = voting_session.minimum_participants - vote_count
        messages.error(
            request,
            f"Для раскрытия не хватает голосов: нужно ещё {remaining}.",
        )
        return redirect(voting_session)

    voting_round.status = VotingRound.Status.REVEALED
    voting_round.revealed_at = timezone.now()
    voting_round.save(update_fields=("status", "revealed_at"))
    return redirect(voting_session)


@login_required
@require_POST
def session_revote(request, pk):
    voting_session = _session_for_user(request.user, pk)
    voting_round = voting_session.current_round
    if not voting_round or voting_round.status != VotingRound.Status.REVEALED:
        messages.error(request, "Повторное голосование сейчас недоступно.")
        return redirect(voting_session)

    with transaction.atomic():
        voting_round.status = VotingRound.Status.CANCELLED
        voting_round.save(update_fields=("status",))
        queue_item = voting_session.queue_items.get(task=voting_round.task)
        _create_round_for_item(voting_session, queue_item)
    messages.info(request, "Начат новый раунд. Предыдущие голоса сохранены в истории.")
    return redirect(voting_session)


@login_required
@require_POST
def session_accept(request, pk):
    voting_session = _session_for_user(request.user, pk)
    voting_round = voting_session.current_round
    if not voting_round or voting_round.status != VotingRound.Status.REVEALED:
        messages.error(request, "Сначала раскройте голоса.")
        return redirect(voting_session)

    try:
        with transaction.atomic():
            queue_item = voting_session.queue_items.get(task=voting_round.task)
            voting_round.task.capture_estimate(voting_round)
            voting_round.status = VotingRound.Status.CLOSED
            voting_round.closed_at = timezone.now()
            voting_round.save(update_fields=("status", "closed_at"))
            queue_item.status = VotingSessionTask.Status.COMPLETED
            queue_item.completed_at = timezone.now()
            queue_item.save(update_fields=("status", "completed_at"))
            next_item = _focus_next_open_item(voting_session, queue_item.position)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect(voting_session)

    messages.success(
        request,
        f"Оценка задачи {voting_round.task.number} сохранена: "
        f"{voting_round.task.estimate_display}."
        + (
            f" Следующая задача для проверки: {next_item.task.number}."
            if next_item
            else " Очередь завершена."
        ),
    )
    return redirect(voting_session)


@login_required
@require_POST
def session_finish(request, pk):
    voting_session = _session_for_user(request.user, pk)
    with transaction.atomic():
        voting_session.rounds.filter(
            status__in=(VotingRound.Status.VOTING, VotingRound.Status.REVEALED)
        ).update(status=VotingRound.Status.CANCELLED)
        voting_session.queue_items.filter(
            status=VotingSessionTask.Status.ACTIVE
        ).update(status=VotingSessionTask.Status.SKIPPED)
        voting_session.current_task = None
        voting_session.status = VotingSession.Status.FINISHED
        voting_session.finished_at = timezone.now()
        voting_session.save(
            update_fields=("current_task", "status", "finished_at")
        )
    messages.success(request, "Сессия голосования завершена.")
    return redirect(voting_session)


@login_required
@require_POST
def session_archive(request, pk):
    voting_session = _session_for_user(request.user, pk)
    if voting_session.status != VotingSession.Status.FINISHED:
        messages.error(request, "В архив можно переместить только завершённую комнату.")
        return redirect(voting_session)
    if voting_session.archived_at is None:
        voting_session.archived_at = timezone.now()
        voting_session.save(update_fields=("archived_at",))
        messages.success(request, f"Комната «{voting_session.name}» перемещена в архив.")
    return redirect(voting_session.project)


@login_required
@require_POST
def session_restore(request, pk):
    voting_session = _session_for_user(request.user, pk)
    if voting_session.archived_at is not None:
        voting_session.archived_at = None
        voting_session.save(update_fields=("archived_at",))
        messages.success(request, f"Комната «{voting_session.name}» восстановлена.")
    return redirect(voting_session)


@login_required
@require_POST
def session_delete(request, pk):
    voting_session = _session_for_user(request.user, pk)
    project = voting_session.project
    session_name = voting_session.name
    voting_session.delete()
    messages.success(request, f"Комната «{session_name}» удалена.")
    return redirect(project)


@login_required
@require_POST
def session_copy(request, pk):
    source = _session_for_user(request.user, pk)
    with transaction.atomic():
        copied = VotingSession.objects.create(
            project=source.project,
            name=_copy_name(source.name),
            minimum_participants=source.minimum_participants,
        )
        VotingSessionTask.objects.bulk_create(
            [
                VotingSessionTask(
                    session=copied,
                    task=item.task,
                    position=item.position,
                )
                for item in source.queue_items.select_related("task")
            ]
        )
    messages.success(
        request,
        f"Создана копия комнаты «{copied.name}» без участников и голосов.",
    )
    return redirect(copied)


@login_required
@require_POST
def participant_resume_rotate(request, pk, participant_pk):
    voting_session = _session_for_user(request.user, pk)
    participant = get_object_or_404(
        Participant, pk=participant_pk, session=voting_session
    )
    participant.client_token = uuid.uuid4()
    participant.save(update_fields=("client_token",))
    return JsonResponse(
        {
            "ok": True,
            "resume_url": _participant_resume_url(
                request, voting_session, participant
            ),
        }
    )


@login_required
@require_GET
def session_state(request, pk):
    voting_session = _session_for_user(request.user, pk)
    voting_round = voting_session.current_round
    voted_ids = set()
    votes = []
    summary = None

    if voting_round:
        round_votes = list(
            voting_round.votes.select_related("participant").order_by(
                "participant__joined_at"
            )
        )
        voted_ids = {vote.participant_id for vote in round_votes}
        if voting_round.status == VotingRound.Status.REVEALED:
            votes = [
                {"name": vote.participant.name, "value": vote.value}
                for vote in round_votes
            ]
            round_summary = voting_round.summary()
            summary = {
                "count": round_summary["count"],
                "sum": round_summary["sum"],
                "average": _format_decimal(round_summary["average"]),
            }

    participant_progress = _participant_progress(voting_session, voted_ids)
    participants = [
        {
            "name": participant.name,
            "current_round_voted": participant.current_round_voted,
            "progress_voted": participant.progress_voted,
            "progress_total": participant.progress_total,
            "progress_status": participant.progress_status,
            "progress_label": participant.progress_label,
            "resume_url": _participant_resume_url(
                request, voting_session, participant
            ),
            "rotate_url": _participant_rotate_url(
                voting_session, participant
            ),
        }
        for participant in participant_progress
    ]
    queue = _queue_summary(voting_session)
    queue_items = list(
        voting_session.queue_items.annotate(
            vote_count=Count("current_round__votes")
        ).values("id", "status", "vote_count")
    )
    minimum_reached = len(voted_ids) >= voting_session.minimum_participants
    return JsonResponse(
        {
            "session_status": voting_session.status,
            "current_task": {
                "number": voting_session.current_task.number,
                "title": voting_session.current_task.title,
                "competency": voting_session.current_task.competency,
                "competency_label": voting_session.current_task.get_competency_display(),
            }
            if voting_session.current_task
            else None,
            "round_status": voting_round.status if voting_round else None,
            "participants": participants,
            "voted_count": len(voted_ids),
            "participant_count": len(participants),
            "completed_participant_count": sum(
                item.progress_status == "completed"
                for item in participant_progress
            ),
            "minimum_participants": voting_session.minimum_participants,
            "minimum_reached": minimum_reached,
            "votes_remaining": max(
                voting_session.minimum_participants - len(voted_ids), 0
            ),
            "queue": {
                "total": queue["total"],
                "completed": queue["completed"],
                "pending": queue["pending"],
                "current_position": queue["current_position"],
            },
            "queue_items": queue_items,
            "votes": votes,
            "summary": summary,
        }
    )


def room(request, token):
    voting_session = get_object_or_404(
        VotingSession.objects.select_related("project", "current_task"),
        public_token=token,
    )
    participant = _participant_from_request(request, voting_session)
    if not participant:
        return render(
            request,
            "poker/room_join.html",
            {"voting_session": voting_session, "join_form": JoinRoomForm()},
        )
    if participant.completed_at:
        return render(
            request,
            "poker/room_completed.html",
            {"voting_session": voting_session},
        )
    return render(
        request,
        "poker/room.html",
        {
            "voting_session": voting_session,
            "participant": participant,
            "estimation_guide": ESTIMATION_GUIDE,
            "resume_url": _participant_resume_url(
                request, voting_session, participant
            ),
        },
    )


@require_POST
def room_join(request, token):
    voting_session = get_object_or_404(VotingSession, public_token=token)
    if voting_session.status == VotingSession.Status.FINISHED:
        messages.error(request, "Эта сессия уже завершена.")
        return redirect("poker:room", token=token)

    form = JoinRoomForm(request.POST)
    if form.is_valid():
        name = form.cleaned_data["name"].strip()
        existing_names = voting_session.participants.values_list("name", flat=True)
        if any(existing.casefold() == name.casefold() for existing in existing_names):
            form.add_error(
                "name",
                "Участник с таким именем уже существует. Используйте "
                "персональную ссылку продолжения или запросите её у организатора.",
            )
        else:
            first_task_id = (
                voting_session.queue_items.filter(
                    status=VotingSessionTask.Status.ACTIVE
                )
                .values_list("task_id", flat=True)
                .first()
                or voting_session.queue_items.values_list(
                    "task_id", flat=True
                ).first()
            )
            participant = Participant.objects.create(
                session=voting_session, name=name, current_task_id=first_task_id
            )
            request.session[_participant_session_key(voting_session)] = str(
                participant.client_token
            )
            request.session.modified = True
            return redirect("poker:room", token=token)

    return render(
        request,
        "poker/room_join.html",
        {"voting_session": voting_session, "join_form": form},
        status=400,
    )


@never_cache
@require_GET
def room_resume(request, token, participant_token):
    voting_session = get_object_or_404(VotingSession, public_token=token)
    participant = voting_session.participants.filter(
        client_token=participant_token
    ).first()
    session_key = _participant_session_key(voting_session)
    if participant is None:
        request.session.pop(session_key, None)
        request.session.modified = True
        messages.error(
            request,
            "Персональная ссылка недействительна. Запросите новую у организатора.",
        )
        return redirect("poker:room", token=token)

    request.session[session_key] = str(participant.client_token)
    request.session.modified = True
    return redirect("poker:room", token=token)


@require_GET
def room_state(request, token):
    voting_session = get_object_or_404(
        VotingSession.objects.select_related("current_task"), public_token=token
    )
    participant = _participant_from_request(request, voting_session)
    if not participant:
        return JsonResponse({"error": "participant_required"}, status=403)

    Participant.objects.filter(pk=participant.pk).update(last_seen_at=timezone.now())
    queue_item = _participant_queue_item(voting_session, participant)
    queue = _participant_queue_summary(voting_session, participant, queue_item)
    response = {
        "session_status": voting_session.status,
        "session_name": voting_session.name,
        "participant_completed": participant.completed_at is not None,
        "current_task": None,
        "round": None,
        "minimum_participants": voting_session.minimum_participants,
        "queue": queue,
    }
    if (
        voting_session.status == VotingSession.Status.DRAFT
        or queue_item is None
        or queue_item.current_round_id is None
    ):
        return JsonResponse(response)

    voting_round = queue_item.current_round
    vote = voting_round.votes.filter(participant=participant).first()
    vote_count = voting_round.votes.count()
    round_data = {
        "status": voting_round.status,
        "number": voting_round.number,
        "has_voted": vote is not None,
        "my_vote": vote.value if vote else None,
        "voted_count": vote_count,
        "participant_count": voting_session.participants.count(),
        "minimum_participants": voting_session.minimum_participants,
        "minimum_reached": vote_count >= voting_session.minimum_participants,
        "votes": [],
        "average": None,
    }
    if voting_round.status in (
        VotingRound.Status.REVEALED,
        VotingRound.Status.CLOSED,
    ):
        revealed_votes = voting_round.votes.select_related("participant").order_by(
            "participant__joined_at"
        )
        round_data["votes"] = [
            {"name": item.participant.name, "value": item.value}
            for item in revealed_votes
        ]
        round_data["average"] = _format_decimal(
            voting_round.summary()["average"]
        )

    response["current_task"] = {
        "id": queue_item.task_id,
        "number": queue_item.task.number,
        "title": queue_item.task.title,
        "competency": queue_item.task.competency,
        "competency_label": queue_item.task.get_competency_display(),
    }
    response["round"] = round_data
    return JsonResponse(response)


@require_POST
def room_vote(request, token):
    voting_session = get_object_or_404(VotingSession, public_token=token)
    participant = _participant_from_request(request, voting_session)
    if not participant:
        return JsonResponse({"error": "participant_required"}, status=403)
    if participant.completed_at:
        return JsonResponse({"error": "participant_completed"}, status=409)
    queue_item = _participant_queue_item(voting_session, participant)
    voting_round = queue_item.current_round if queue_item else None
    if not voting_round or voting_round.status != VotingRound.Status.VOTING:
        return JsonResponse({"error": "voting_closed"}, status=409)

    try:
        value = int(request.POST.get("value", ""))
    except (TypeError, ValueError):
        return JsonResponse({"error": "invalid_value"}, status=400)
    if value not in ESTIMATION_VALUES:
        return JsonResponse({"error": "invalid_value"}, status=400)

    Vote.objects.update_or_create(
        voting_round=voting_round,
        participant=participant,
        defaults={"value": value},
    )
    return JsonResponse({"ok": True, "value": value})


@require_POST
def room_navigate(request, token):
    voting_session = get_object_or_404(VotingSession, public_token=token)
    participant = _participant_from_request(request, voting_session)
    if not participant:
        return JsonResponse({"error": "participant_required"}, status=403)
    if participant.completed_at:
        return JsonResponse({"error": "participant_completed"}, status=409)

    direction = request.POST.get("direction")
    if direction not in ("previous", "next", "missing"):
        return JsonResponse({"error": "invalid_direction"}, status=400)

    queue_item = _participant_queue_item(voting_session, participant)
    if queue_item is None:
        return JsonResponse({"error": "empty_queue"}, status=409)

    if direction == "missing":
        queue = _participant_queue_summary(
            voting_session, participant, queue_item
        )
        target = (
            voting_session.queue_items.select_related("task")
            .filter(task_id=queue["first_missing_task_id"])
            .first()
            if queue["first_missing_task_id"]
            else None
        )
        if target is None:
            return JsonResponse({"error": "no_missing_tasks"}, status=409)
    else:
        position_filter = (
            {"position__gt": queue_item.position}
            if direction == "next"
            else {"position__lt": queue_item.position}
        )
        target_items = voting_session.queue_items.select_related("task").filter(
            **position_filter
        )
        target = (
            target_items.first()
            if direction == "next"
            else target_items.last()
        )
    if target is None:
        return JsonResponse({"error": "queue_boundary"}, status=409)

    participant.current_task = target.task
    participant.save(update_fields=("current_task",))
    return JsonResponse(
        {"ok": True, "position": target.position, "task_id": target.task_id}
    )


@require_POST
def room_complete(request, token):
    voting_session = get_object_or_404(VotingSession, public_token=token)
    participant = _participant_from_request(request, voting_session)
    if not participant:
        return JsonResponse({"error": "participant_required"}, status=403)
    queue_item = _participant_queue_item(voting_session, participant)
    if queue_item is None:
        return JsonResponse({"error": "empty_queue"}, status=409)
    queue = _participant_queue_summary(voting_session, participant, queue_item)
    if participant.completed_at and queue["all_voted"]:
        return JsonResponse({"ok": True})
    if (
        voting_session.status != VotingSession.Status.ACTIVE
        or queue_item.current_round_id is None
    ):
        return JsonResponse({"error": "voting_not_started"}, status=409)
    if voting_session.queue_items.filter(
        position__gt=queue_item.position
    ).exists():
        return JsonResponse({"error": "last_task_required"}, status=409)
    if not queue["all_voted"]:
        return JsonResponse(
            {
                "error": "incomplete_tasks",
                "voted": queue["voted"],
                "total": queue["total"],
                "missing": queue["missing"],
                "first_missing_task_id": queue["first_missing_task_id"],
                "first_missing_position": queue["first_missing_position"],
            },
            status=409,
        )

    participant.completed_at = timezone.now()
    participant.save(update_fields=("completed_at",))
    return JsonResponse({"ok": True})


@login_required
@require_POST
def sprint_create(request, pk):
    project = _project_for_user(request.user, pk)
    form = SprintForm(request.POST)
    if form.is_valid():
        sprint = form.save(commit=False)
        sprint.project = project
        sprint.save()
        messages.success(request, f"Спринт «{sprint.name}» создан.")
        return redirect(sprint)
    messages.error(request, "Не удалось создать спринт: проверьте параметры.")
    return redirect(project)


@login_required
def sprint_detail(request, pk):
    sprint = _sprint_for_user(request.user, pk)
    sprint_items = sprint.sprint_tasks.select_related("task", "transferred_to")
    planned_items = sprint_items.filter(status=SprintTask.Status.PLANNED)
    available_tasks = (
        sprint.project.tasks.filter(
            status=Task.Status.ESTIMATED,
            completed_at__isnull=True,
        )
        .exclude(sprint_items__sprint=sprint)
        .exclude(sprint_items__status=SprintTask.Status.PLANNED)
        .distinct()
    )
    transfer_targets = sprint.project.sprints.filter(
        archived_at__isnull=True,
        status__in=(Sprint.Status.PLANNING, Sprint.Status.ACTIVE),
    ).exclude(pk=sprint.pk)
    return render(
        request,
        "poker/sprint_detail.html",
        {
            "sprint": sprint,
            "sprint_items": sprint_items,
            "planned_items": planned_items,
            "available_tasks": available_tasks,
            "transfer_targets": transfer_targets,
        },
    )


@login_required
@require_POST
def sprint_add_tasks(request, pk):
    sprint = _sprint_for_user(request.user, pk)
    if sprint.archived_at is not None or sprint.status == Sprint.Status.COMPLETED:
        messages.error(
            request,
            "Нельзя менять состав завершённого или архивного спринта.",
        )
        return redirect(sprint)
    task_ids = request.POST.getlist("task_ids")
    tasks = sprint.project.tasks.filter(
        pk__in=task_ids,
        status=Task.Status.ESTIMATED,
        completed_at__isnull=True,
    ).exclude(sprint_items__sprint=sprint).exclude(
        sprint_items__status=SprintTask.Status.PLANNED
    ).distinct()
    position = (
        sprint.sprint_tasks.aggregate(value=Max("position"))["value"] or 0
    )
    created = 0
    with transaction.atomic():
        for task in tasks:
            position += 1
            SprintTask.objects.create(
                sprint=sprint,
                task=task,
                position=position,
                status=SprintTask.Status.PLANNED,
            )
            created += 1
    messages.success(request, f"В спринт добавлено задач: {created}.")
    _warn_if_over_capacity(request, sprint)
    return redirect(sprint)


@login_required
@require_POST
def sprint_remove_task(request, pk, task_pk):
    sprint = _sprint_for_user(request.user, pk)
    if sprint.archived_at is not None or sprint.status == Sprint.Status.COMPLETED:
        messages.error(
            request,
            "Нельзя менять состав завершённого или архивного спринта.",
        )
        return redirect(sprint)
    item = get_object_or_404(
        SprintTask,
        sprint=sprint,
        task_id=task_pk,
        status=SprintTask.Status.PLANNED,
    )
    item.delete()
    messages.info(request, "Задача удалена из спринта.")
    return redirect(sprint)


@login_required
@require_POST
def sprint_set_status(request, pk):
    sprint = _sprint_for_user(request.user, pk)
    if sprint.archived_at is not None:
        messages.error(request, "Сначала восстановите спринт из архива.")
        return redirect(sprint)

    status = request.POST.get("status")
    valid_statuses = {value for value, _label in Sprint.Status.choices}
    if status not in valid_statuses:
        messages.error(request, "Неизвестный статус спринта.")
        return redirect(sprint)

    sprint.status = status
    sprint.save(update_fields=("status",))
    messages.success(
        request,
        f"Статус спринта изменён: {sprint.get_status_display()}.",
    )
    return redirect(sprint)


@login_required
@require_POST
def sprint_archive(request, pk):
    sprint = _sprint_for_user(request.user, pk)
    if sprint.status != Sprint.Status.COMPLETED:
        messages.error(request, "В архив можно переместить только завершённый спринт.")
        return redirect(sprint)
    if sprint.archived_at is None:
        sprint.archived_at = timezone.now()
        sprint.save(update_fields=("archived_at",))
        messages.success(request, f"Спринт «{sprint.name}» перемещён в архив.")
    return redirect(sprint.project)


@login_required
@require_POST
def sprint_restore(request, pk):
    sprint = _sprint_for_user(request.user, pk)
    if sprint.archived_at is not None:
        sprint.archived_at = None
        sprint.save(update_fields=("archived_at",))
        messages.success(request, f"Спринт «{sprint.name}» восстановлен.")
    return redirect(sprint)


@login_required
@require_POST
def sprint_delete(request, pk):
    sprint = _sprint_for_user(request.user, pk)
    project = sprint.project
    sprint_name = sprint.name
    sprint.delete()
    messages.success(request, f"Спринт «{sprint_name}» удалён.")
    return redirect(project)


@login_required
@require_POST
def sprint_copy(request, pk):
    source = _sprint_for_user(request.user, pk)
    with transaction.atomic():
        copied = Sprint.objects.create(
            project=source.project,
            name=_copy_name(source.name),
            status=Sprint.Status.PLANNING,
            goal=source.goal,
            capacity=source.capacity,
        )
        SprintTask.objects.bulk_create(
            [
                SprintTask(
                    sprint=copied,
                    task=item.task,
                    position=item.position,
                    status=SprintTask.Status.PLANNED,
                )
                for item in source.sprint_tasks.select_related("task").filter(
                    status=SprintTask.Status.PLANNED
                )
            ]
        )
    messages.success(
        request,
        f"Создан планируемый спринт «{copied.name}» без дат.",
    )
    _warn_if_over_capacity(request, copied)
    return redirect(copied)


@login_required
@require_POST
def sprint_transfer_tasks(request, pk):
    source = _sprint_for_user(request.user, pk)
    if source.archived_at is not None:
        messages.error(request, "Нельзя переносить задачи из архивного спринта.")
        return redirect(source)

    target = get_object_or_404(
        Sprint,
        pk=request.POST.get("target_sprint"),
        project=source.project,
        archived_at__isnull=True,
        status__in=(Sprint.Status.PLANNING, Sprint.Status.ACTIVE),
    )
    if target.pk == source.pk:
        messages.error(request, "Выберите другой спринт.")
        return redirect(source)

    task_ids = request.POST.getlist("task_ids")
    source_items = list(
        source.sprint_tasks.select_related("task").filter(
            task_id__in=task_ids,
            status=SprintTask.Status.PLANNED,
        )
    )
    position = target.sprint_tasks.aggregate(value=Max("position"))["value"] or 0
    transferred = 0
    now = timezone.now()
    with transaction.atomic():
        for source_item in source_items:
            target_item = target.sprint_tasks.filter(task=source_item.task).first()
            if target_item is None:
                position += 1
                SprintTask.objects.create(
                    sprint=target,
                    task=source_item.task,
                    position=position,
                    status=SprintTask.Status.PLANNED,
                )
            elif target_item.status != SprintTask.Status.PLANNED:
                target_item.status = SprintTask.Status.PLANNED
                target_item.transferred_to = None
                target_item.transferred_at = None
                target_item.save(
                    update_fields=("status", "transferred_to", "transferred_at")
                )

            source_item.status = SprintTask.Status.TRANSFERRED
            source_item.transferred_to = target
            source_item.transferred_at = now
            source_item.save(
                update_fields=("status", "transferred_to", "transferred_at")
            )
            transferred += 1

    if transferred:
        messages.success(
            request,
            f"В спринт «{target.name}» перенесено задач: {transferred}.",
        )
        _warn_if_over_capacity(request, target)
    else:
        messages.info(request, "Выберите задачи для переноса.")
    return redirect(source)


@login_required
def sprint_export(request, pk):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    sprint = _sprint_for_user(request.user, pk)
    items = list(
        sprint.sprint_tasks.select_related("task").filter(
            status=SprintTask.Status.PLANNED
        )
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Задачи спринта"
    headers = (
        "№",
        "Номер задачи",
        "Название",
        "Средняя оценка",
        "Сумма голосов",
        "Количество голосов",
        "Тип задачи",
    )
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="243B53")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_number, item in enumerate(items, start=2):
        task = item.task
        sheet.append(
            (
                row_number - 1,
                task.number,
                task.title,
                f"=E{row_number}/F{row_number}" if task.estimate_count else None,
                task.estimate_sum,
                task.estimate_count,
                task.get_competency_display(),
            )
        )
        sheet.cell(row=row_number, column=4).number_format = "0.00"

    if items:
        total_row = len(items) + 2
        sheet.cell(row=total_row, column=3, value="Итого")
        sheet.cell(row=total_row, column=3).font = Font(bold=True)
        sheet.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row - 1})")
        sheet.cell(row=total_row, column=4).font = Font(bold=True)
        sheet.cell(row=total_row, column=4).number_format = "0.00"

    widths = (6, 20, 70, 20, 18, 22, 20)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{max(1, len(items) + 1)}"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    safe_name = "".join(
        character if character.isalnum() or character in "-_" else "_"
        for character in sprint.name
    )
    response["Content-Disposition"] = (
        f'attachment; filename="sprint_{safe_name or sprint.pk}.xlsx"'
    )
    workbook.save(response)
    return response

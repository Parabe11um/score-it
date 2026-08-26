import uuid
from io import BytesIO
from datetime import timedelta
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import Client, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from poker.models import (
    Participant,
    Project,
    Sprint,
    SprintTask,
    Task,
    Vote,
    VotingRound,
    VotingSession,
    VotingSessionTask,
)


class OrganizerViewsTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user("owner", password="secret")
        self.other_user = get_user_model().objects.create_user(
            "other", password="secret"
        )
        self.project = Project.objects.create(owner=self.user, name="ABS Core")

    def test_dashboard_requires_login(self):
        response = self.client.get(reverse("poker:dashboard"))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("login"), response.url)

    def test_bulk_import_creates_and_updates_tasks(self):
        self.client.force_login(self.user)
        url = reverse("poker:task_import", args=[self.project.pk])

        response = self.client.post(
            url,
            {"tasks_text": "ABS-1 | Первая задача\nABS-2 Вторая задача"},
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.project.tasks.count(), 2)

        self.client.post(url, {"tasks_text": "ABS-1 | Обновлённое название"})
        self.assertEqual(self.project.tasks.count(), 2)
        self.assertEqual(
            self.project.tasks.get(number="ABS-1").title, "Обновлённое название"
        )

    def test_bulk_import_supports_default_and_per_line_competencies(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("poker:task_import", args=[self.project.pk]),
            {
                "competency": Task.Competency.DEVELOPMENT,
                "tasks_text": (
                    "[Аналитика] ABS-1 | Описать требования\n"
                    "ABS-2 | Реализовать обработчик\n"
                    "[QA] ABS-3 | Проверить обработчик"
                ),
            },
        )

        self.assertRedirects(response, self.project.get_absolute_url())
        self.assertEqual(
            self.project.tasks.get(number="ABS-1").competency,
            Task.Competency.ANALYSIS,
        )
        self.assertEqual(
            self.project.tasks.get(number="ABS-2").competency,
            Task.Competency.DEVELOPMENT,
        )
        self.assertEqual(
            self.project.tasks.get(number="ABS-3").competency,
            Task.Competency.TESTING,
        )

    def test_bulk_import_rejects_unknown_competency_prefix(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("poker:task_import", args=[self.project.pk]),
            {"tasks_text": "[Дизайн] ABS-1 | Нарисовать экран"},
        )

        self.assertEqual(response.status_code, 400)
        self.assertContains(response, "Не удалось распознать строки: 1", status_code=400)
        self.assertFalse(self.project.tasks.exists())

    def test_other_organizer_cannot_open_project(self):
        self.client.force_login(self.other_user)
        response = self.client.get(self.project.get_absolute_url())
        self.assertEqual(response.status_code, 404)

    def test_project_page_offers_bulk_queue_and_minimum_votes(self):
        Task.objects.create(project=self.project, number="ABS-1", title="Первая")
        self.client.force_login(self.user)

        response = self.client.get(self.project.get_absolute_url())

        self.assertContains(response, "Создать комнату с очередью")
        self.assertContains(response, "Минимум голосов")
        self.assertContains(response, "checked")

    def test_sprint_creation_accepts_competency_capacities(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("poker:sprint_create", args=[self.project.pk]),
            {
                "name": "Спринт по компетенциям",
                "goal": "Проверить лимиты команды",
                "analysis_capacity": "24",
                "development_capacity": "60.5",
                "testing_capacity": "32",
            },
        )

        sprint = Sprint.objects.get(name="Спринт по компетенциям")
        self.assertRedirects(response, sprint.get_absolute_url())
        self.assertEqual(sprint.analysis_capacity, Decimal("24"))
        self.assertEqual(sprint.development_capacity, Decimal("60.5"))
        self.assertEqual(sprint.testing_capacity, Decimal("32"))
        self.assertIsNone(sprint.capacity)
        self.assertEqual(sprint.capacity_total, Decimal("116.5"))

    def test_session_creation_adds_selected_tasks_to_ordered_queue(self):
        first = Task.objects.create(
            project=self.project, number="ABS-1", title="Первая"
        )
        second = Task.objects.create(
            project=self.project, number="ABS-2", title="Вторая"
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("poker:session_create", args=[self.project.pk]),
            {
                "name": "Пакетная оценка",
                "minimum_participants": 3,
                "task_ids": [first.pk, second.pk],
            },
        )

        voting_session = VotingSession.objects.get(name="Пакетная оценка")
        self.assertRedirects(response, voting_session.get_absolute_url())
        self.assertEqual(voting_session.minimum_participants, 3)
        self.assertEqual(
            list(
                voting_session.queue_items.values_list("task_id", "position")
            ),
            [(first.pk, 1), (second.pk, 2)],
        )

    def test_task_filters_cover_new_estimated_sprint_and_completed_tasks(self):
        new_task = Task.objects.create(
            project=self.project, number="ABS-1", title="Новая"
        )
        queued_task = Task.objects.create(
            project=self.project, number="ABS-2", title="В очереди"
        )
        estimated_task = Task.objects.create(
            project=self.project,
            number="ABS-3",
            title="Оценена",
            status=Task.Status.ESTIMATED,
            estimate_sum=8,
            estimate_count=1,
        )
        sprint_task = Task.objects.create(
            project=self.project,
            number="ABS-4",
            title="В спринте",
            status=Task.Status.ESTIMATED,
            estimate_sum=12,
            estimate_count=1,
        )
        completed_task = Task.objects.create(
            project=self.project,
            number="ABS-5",
            title="Завершена",
            completed_at=timezone.now(),
        )
        voting_session = VotingSession.objects.create(
            project=self.project, name="Комната"
        )
        VotingSessionTask.objects.create(
            session=voting_session, task=queued_task, position=1
        )
        sprint = Sprint.objects.create(project=self.project, name="Спринт")
        SprintTask.objects.create(sprint=sprint, task=sprint_task, position=1)
        self.client.force_login(self.user)

        expected = {
            "new": {new_task.pk},
            "unestimated": {new_task.pk, queued_task.pk},
            "estimated": {estimated_task.pk, sprint_task.pk},
            "sprint": {sprint_task.pk},
            "completed": {completed_task.pk},
        }
        for task_filter, expected_ids in expected.items():
            response = self.client.get(
                self.project.get_absolute_url(), {"tasks": task_filter}
            )
            self.assertEqual(response.status_code, 200)
            self.assertEqual(
                set(response.context["tasks"].values_list("pk", flat=True)),
                expected_ids,
            )

    def test_task_competency_filters_and_inline_update(self):
        analysis = Task.objects.create(
            project=self.project,
            number="ABS-A",
            title="Аналитическая задача",
            competency=Task.Competency.ANALYSIS,
        )
        development = Task.objects.create(
            project=self.project,
            number="ABS-D",
            title="Задача разработки",
            competency=Task.Competency.DEVELOPMENT,
        )
        Task.objects.create(
            project=self.project,
            number="ABS-N",
            title="Без типа",
        )
        self.client.force_login(self.user)

        response = self.client.get(
            self.project.get_absolute_url(),
            {"competency": "analysis"},
        )
        self.assertEqual(
            list(response.context["tasks"].values_list("pk", flat=True)),
            [analysis.pk],
        )
        self.assertContains(response, "Все направления")
        self.assertContains(response, "Аналитика")

        update = self.client.post(
            reverse(
                "poker:task_competency_update",
                args=[self.project.pk, development.pk],
            ),
            {"competency": Task.Competency.TESTING},
        )
        self.assertRedirects(update, self.project.get_absolute_url())
        development.refresh_from_db()
        self.assertEqual(development.competency, Task.Competency.TESTING)

    def test_organizer_can_complete_reopen_and_delete_task(self):
        task = Task.objects.create(
            project=self.project, number="ABS-1", title="Задача"
        )
        self.client.force_login(self.user)

        self.client.post(
            reverse("poker:task_complete", args=[self.project.pk, task.pk])
        )
        task.refresh_from_db()
        self.assertIsNotNone(task.completed_at)

        self.client.post(
            reverse("poker:task_reopen", args=[self.project.pk, task.pk])
        )
        task.refresh_from_db()
        self.assertIsNone(task.completed_at)

        self.client.post(
            reverse("poker:task_delete", args=[self.project.pk, task.pk])
        )
        self.assertFalse(Task.objects.filter(pk=task.pk).exists())

    def test_other_organizer_cannot_delete_project_entities(self):
        task = Task.objects.create(
            project=self.project, number="ABS-1", title="Задача"
        )
        voting_session = VotingSession.objects.create(
            project=self.project, name="Комната"
        )
        sprint = Sprint.objects.create(project=self.project, name="Спринт")
        self.client.force_login(self.other_user)

        responses = (
            self.client.post(
                reverse("poker:task_delete", args=[self.project.pk, task.pk])
            ),
            self.client.post(
                reverse(
                    "poker:task_competency_update",
                    args=[self.project.pk, task.pk],
                ),
                {"competency": Task.Competency.ANALYSIS},
            ),
            self.client.post(reverse("poker:session_delete", args=[voting_session.pk])),
            self.client.post(reverse("poker:sprint_delete", args=[sprint.pk])),
            self.client.post(
                reverse("poker:sprint_capacity_update", args=[sprint.pk]),
                {
                    "analysis_capacity": "8",
                    "development_capacity": "20",
                    "testing_capacity": "12",
                },
            ),
        )
        self.assertTrue(all(response.status_code == 404 for response in responses))
        self.assertTrue(Task.objects.filter(pk=task.pk).exists())
        self.assertTrue(VotingSession.objects.filter(pk=voting_session.pk).exists())
        self.assertTrue(Sprint.objects.filter(pk=sprint.pk).exists())

    def test_finished_room_can_be_archived_restored_and_deleted(self):
        voting_session = VotingSession.objects.create(
            project=self.project,
            name="Старая комната",
            status=VotingSession.Status.FINISHED,
            finished_at=timezone.now(),
        )
        self.client.force_login(self.user)

        self.client.post(reverse("poker:session_archive", args=[voting_session.pk]))
        voting_session.refresh_from_db()
        self.assertIsNotNone(voting_session.archived_at)
        active_page = self.client.get(self.project.get_absolute_url())
        self.assertNotIn(voting_session, list(active_page.context["sessions"]))
        archive_page = self.client.get(
            self.project.get_absolute_url(), {"archive": "1"}
        )
        self.assertIn(voting_session, list(archive_page.context["sessions"]))

        self.client.post(reverse("poker:session_restore", args=[voting_session.pk]))
        voting_session.refresh_from_db()
        self.assertIsNone(voting_session.archived_at)

        self.client.post(reverse("poker:session_delete", args=[voting_session.pk]))
        self.assertFalse(VotingSession.objects.filter(pk=voting_session.pk).exists())

    def test_dashboard_summarizes_active_work(self):
        new_task = Task.objects.create(
            project=self.project, number="ABS-1", title="Новая задача"
        )
        active_room = VotingSession.objects.create(
            project=self.project,
            name="Активная оценка",
            status=VotingSession.Status.ACTIVE,
        )
        Participant.objects.create(
            session=active_room,
            name="Анна",
            completed_at=timezone.now(),
        )
        Participant.objects.create(session=active_room, name="Борис")
        active_sprint = Sprint.objects.create(
            project=self.project,
            name="Текущий спринт",
            status=Sprint.Status.ACTIVE,
            end_date=timezone.localdate() + timedelta(days=5),
            capacity=20,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse("poker:dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["new_task_count"], 1)
        self.assertEqual(response.context["unestimated_task_count"], 1)
        self.assertEqual(response.context["incomplete_participant_count"], 1)
        self.assertEqual(list(response.context["active_rooms"]), [active_room])
        self.assertEqual(list(response.context["active_sprints"]), [active_sprint])
        self.assertContains(response, "Активная оценка")
        self.assertContains(response, "Текущий спринт")


class VotingFlowTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user("owner", password="secret")
        self.project = Project.objects.create(owner=self.user, name="ABS Core")
        self.task = Task.objects.create(
            project=self.project, number="ABS-101", title="Рассчитать комиссию"
        )
        self.voting_session = VotingSession.objects.create(
            project=self.project, name="Оценка спринта"
        )
        self.organizer = Client()
        self.organizer.force_login(self.user)

    def join_participant(self, name):
        client = Client()
        response = client.post(
            reverse("poker:room_join", args=[self.voting_session.public_token]),
            {"name": name},
        )
        self.assertEqual(response.status_code, 302)
        return client

    def test_complete_voting_flow_saves_exact_average(self):
        first = self.join_participant("Анна")
        second = self.join_participant("Борис")

        response = self.organizer.post(
            reverse(
                "poker:session_start_task",
                args=[self.voting_session.pk, self.task.pk],
            )
        )
        self.assertEqual(response.status_code, 302)

        vote_url = reverse(
            "poker:room_vote", args=[self.voting_session.public_token]
        )
        self.assertEqual(first.post(vote_url, {"value": 2}).status_code, 200)
        self.assertEqual(second.post(vote_url, {"value": 8}).status_code, 200)

        self.organizer.post(
            reverse("poker:session_reveal", args=[self.voting_session.pk])
        )
        state = first.get(
            reverse("poker:room_state", args=[self.voting_session.public_token])
        ).json()
        self.assertEqual(state["round"]["status"], VotingRound.Status.REVEALED)
        self.assertEqual(state["round"]["average"], "5")
        self.assertEqual(len(state["round"]["votes"]), 2)

        self.organizer.post(
            reverse("poker:session_accept", args=[self.voting_session.pk])
        )
        self.task.refresh_from_db()
        self.voting_session.refresh_from_db()

        self.assertEqual(self.task.estimate_sum, 10)
        self.assertEqual(self.task.estimate_count, 2)
        self.assertEqual(self.task.estimate_display, "5")
        self.assertIsNone(self.voting_session.current_task)

    def test_vote_rejects_value_outside_scale(self):
        participant = self.join_participant("Анна")
        self.organizer.post(
            reverse(
                "poker:session_start_task",
                args=[self.voting_session.pk, self.task.pk],
            )
        )
        response = participant.post(
            reverse("poker:room_vote", args=[self.voting_session.public_token]),
            {"value": 5},
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(Vote.objects.count(), 0)

    def test_participant_room_explains_estimation_scale(self):
        participant = self.join_participant("Анна")

        response = participant.get(
            reverse("poker:room", args=[self.voting_session.public_token])
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Как оценивать?")
        self.assertContains(response, "Оценивайте весь путь задачи")
        self.assertContains(response, "Не выбирайте 0")
        for label in (
            "Нет работы",
            "Минимальная",
            "Небольшая",
            "Обычная",
            "Средняя",
            "Большая",
            "Разбить",
            "Эпик",
        ):
            self.assertContains(response, label)

    def test_reveal_waits_for_configured_minimum_votes(self):
        first = self.join_participant("Анна")
        self.organizer.post(
            reverse(
                "poker:session_start_task",
                args=[self.voting_session.pk, self.task.pk],
            )
        )
        first.post(
            reverse("poker:room_vote", args=[self.voting_session.public_token]),
            {"value": 8},
        )

        self.organizer.post(
            reverse("poker:session_reveal", args=[self.voting_session.pk])
        )

        self.voting_session.refresh_from_db()
        voting_round = self.voting_session.current_round
        self.assertEqual(voting_round.status, VotingRound.Status.VOTING)

    def test_accepting_estimate_starts_next_queued_task(self):
        second_task = Task.objects.create(
            project=self.project, number="ABS-102", title="Следующая задача"
        )
        VotingSessionTask.objects.bulk_create(
            [
                VotingSessionTask(
                    session=self.voting_session, task=self.task, position=1
                ),
                VotingSessionTask(
                    session=self.voting_session, task=second_task, position=2
                ),
            ]
        )
        first = self.join_participant("Анна")
        second = self.join_participant("Борис")
        self.organizer.post(
            reverse("poker:session_start", args=[self.voting_session.pk])
        )
        vote_url = reverse(
            "poker:room_vote", args=[self.voting_session.public_token]
        )
        first.post(vote_url, {"value": 4})
        second.post(vote_url, {"value": 8})
        self.organizer.post(
            reverse("poker:session_reveal", args=[self.voting_session.pk])
        )

        self.organizer.post(
            reverse("poker:session_accept", args=[self.voting_session.pk])
        )

        self.voting_session.refresh_from_db()
        self.assertEqual(self.voting_session.current_task, second_task)
        self.assertEqual(
            self.voting_session.queue_items.get(task=self.task).status,
            VotingSessionTask.Status.COMPLETED,
        )
        self.assertEqual(
            self.voting_session.queue_items.get(task=second_task).status,
            VotingSessionTask.Status.ACTIVE,
        )
        self.assertEqual(
            self.voting_session.current_round.status, VotingRound.Status.VOTING
        )

    def test_start_opens_every_queued_task_for_async_voting(self):
        second_task = Task.objects.create(
            project=self.project, number="ABS-102", title="Вторая задача"
        )
        VotingSessionTask.objects.bulk_create(
            [
                VotingSessionTask(
                    session=self.voting_session, task=self.task, position=1
                ),
                VotingSessionTask(
                    session=self.voting_session, task=second_task, position=2
                ),
            ]
        )

        self.organizer.post(
            reverse("poker:session_start", args=[self.voting_session.pk])
        )

        queue_items = list(
            self.voting_session.queue_items.select_related("current_round")
        )
        self.assertEqual(
            [item.status for item in queue_items],
            [VotingSessionTask.Status.ACTIVE, VotingSessionTask.Status.ACTIVE],
        )
        self.assertTrue(all(item.current_round_id for item in queue_items))
        self.assertEqual(
            {item.current_round.status for item in queue_items},
            {VotingRound.Status.VOTING},
        )

    def test_participants_navigate_and_vote_independently(self):
        second_task = Task.objects.create(
            project=self.project, number="ABS-102", title="Вторая задача"
        )
        VotingSessionTask.objects.bulk_create(
            [
                VotingSessionTask(
                    session=self.voting_session, task=self.task, position=1
                ),
                VotingSessionTask(
                    session=self.voting_session, task=second_task, position=2
                ),
            ]
        )
        anna = self.join_participant("Анна")
        boris = self.join_participant("Борис")
        room_response = anna.get(
            reverse("poker:room", args=[self.voting_session.public_token])
        )
        self.assertContains(room_response, "Вперёд")
        self.organizer.post(
            reverse("poker:session_start", args=[self.voting_session.pk])
        )
        vote_url = reverse(
            "poker:room_vote", args=[self.voting_session.public_token]
        )
        navigate_url = reverse(
            "poker:room_navigate", args=[self.voting_session.public_token]
        )
        state_url = reverse(
            "poker:room_state", args=[self.voting_session.public_token]
        )

        self.assertEqual(anna.post(vote_url, {"value": 2}).status_code, 200)
        self.assertEqual(
            anna.post(navigate_url, {"direction": "next"}).status_code, 200
        )
        self.assertEqual(anna.post(vote_url, {"value": 8}).status_code, 200)
        self.assertEqual(boris.post(vote_url, {"value": 4}).status_code, 200)

        anna_state = anna.get(state_url).json()
        boris_state = boris.get(state_url).json()
        self.assertEqual(anna_state["current_task"]["id"], second_task.pk)
        self.assertEqual(anna_state["round"]["my_vote"], 8)
        self.assertEqual(anna_state["queue"]["total"], 2)
        self.assertEqual(anna_state["queue"]["voted"], 2)
        self.assertEqual(anna_state["queue"]["missing"], 0)
        self.assertTrue(anna_state["queue"]["all_voted"])
        self.assertEqual(boris_state["current_task"]["id"], self.task.pk)
        self.assertEqual(boris_state["round"]["my_vote"], 4)
        self.assertEqual(boris_state["queue"]["voted"], 1)
        self.assertEqual(boris_state["queue"]["missing"], 1)
        self.assertFalse(boris_state["queue"]["all_voted"])

        organizer_state = self.organizer.get(
            reverse("poker:session_state", args=[self.voting_session.pk])
        ).json()
        self.assertEqual(
            [item["vote_count"] for item in organizer_state["queue_items"]],
            [2, 1],
        )

        anna.post(navigate_url, {"direction": "previous"})
        returned_state = anna.get(state_url).json()
        self.assertEqual(returned_state["current_task"]["id"], self.task.pk)
        self.assertEqual(returned_state["round"]["my_vote"], 2)

    def test_participant_finishes_on_last_task_and_sees_thank_you_page(self):
        second_task = Task.objects.create(
            project=self.project, number="ABS-102", title="Вторая задача"
        )
        VotingSessionTask.objects.bulk_create(
            [
                VotingSessionTask(
                    session=self.voting_session, task=self.task, position=1
                ),
                VotingSessionTask(
                    session=self.voting_session, task=second_task, position=2
                ),
            ]
        )
        participant_client = self.join_participant("Анна")
        self.organizer.post(
            reverse("poker:session_start", args=[self.voting_session.pk])
        )
        complete_url = reverse(
            "poker:room_complete", args=[self.voting_session.public_token]
        )
        navigate_url = reverse(
            "poker:room_navigate", args=[self.voting_session.public_token]
        )

        early_response = participant_client.post(complete_url)
        self.assertEqual(early_response.status_code, 409)
        self.assertEqual(early_response.json()["error"], "last_task_required")

        vote_url = reverse(
            "poker:room_vote", args=[self.voting_session.public_token]
        )
        participant_client.post(vote_url, {"value": 4})
        participant_client.post(navigate_url, {"direction": "next"})
        participant_client.post(vote_url, {"value": 8})
        response = participant_client.post(complete_url)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["ok"])

        participant = Participant.objects.get(name="Анна")
        self.assertIsNotNone(participant.completed_at)
        room_response = participant_client.get(
            reverse("poker:room", args=[self.voting_session.public_token])
        )
        self.assertContains(room_response, "Спасибо за оценку!")
        self.assertContains(room_response, "Ваши ответы сохранены.")
        self.assertNotContains(room_response, "Вперёд")

        state_response = participant_client.get(
            reverse("poker:room_state", args=[self.voting_session.public_token])
        )
        self.assertTrue(state_response.json()["participant_completed"])
        self.assertEqual(
            participant_client.post(
                reverse(
                    "poker:room_vote", args=[self.voting_session.public_token]
                ),
                {"value": 8},
            ).status_code,
            409,
        )

    def test_participant_cannot_finish_with_missing_votes_and_can_jump_to_first_missing(self):
        second_task = Task.objects.create(
            project=self.project, number="ABS-102", title="Вторая задача"
        )
        VotingSessionTask.objects.bulk_create(
            [
                VotingSessionTask(
                    session=self.voting_session, task=self.task, position=1
                ),
                VotingSessionTask(
                    session=self.voting_session, task=second_task, position=2
                ),
            ]
        )
        participant_client = self.join_participant("Анна")
        self.organizer.post(
            reverse("poker:session_start", args=[self.voting_session.pk])
        )
        vote_url = reverse(
            "poker:room_vote", args=[self.voting_session.public_token]
        )
        navigate_url = reverse(
            "poker:room_navigate", args=[self.voting_session.public_token]
        )
        complete_url = reverse(
            "poker:room_complete", args=[self.voting_session.public_token]
        )

        participant_client.post(navigate_url, {"direction": "next"})
        participant_client.post(vote_url, {"value": 8})
        response = participant_client.post(complete_url)

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()["error"], "incomplete_tasks")
        self.assertEqual(response.json()["voted"], 1)
        self.assertEqual(response.json()["missing"], 1)
        self.assertEqual(
            response.json()["first_missing_task_id"], self.task.pk
        )
        participant = Participant.objects.get(name="Анна")
        self.assertIsNone(participant.completed_at)

        jump_response = participant_client.post(
            navigate_url, {"direction": "missing"}
        )
        self.assertEqual(jump_response.status_code, 200)
        state = participant_client.get(
            reverse("poker:room_state", args=[self.voting_session.public_token])
        ).json()
        self.assertEqual(state["current_task"]["id"], self.task.pk)
        self.assertFalse(state["queue"]["current_has_voted"])

    def test_adding_task_to_active_room_reopens_completed_participant(self):
        VotingSessionTask.objects.create(
            session=self.voting_session, task=self.task, position=1
        )
        participant_client = self.join_participant("Анна")
        self.organizer.post(
            reverse("poker:session_start", args=[self.voting_session.pk])
        )
        participant_client.post(
            reverse("poker:room_vote", args=[self.voting_session.public_token]),
            {"value": 4},
        )
        participant_client.post(
            reverse("poker:room_complete", args=[self.voting_session.public_token])
        )
        participant = Participant.objects.get(name="Анна")
        self.assertIsNotNone(participant.completed_at)

        second_task = Task.objects.create(
            project=self.project, number="ABS-102", title="Вторая задача"
        )
        self.organizer.post(
            reverse("poker:session_queue_add", args=[self.voting_session.pk]),
            {"task_ids": [second_task.pk]},
        )

        participant.refresh_from_db()
        self.assertIsNone(participant.completed_at)
        state = participant_client.get(
            reverse("poker:room_state", args=[self.voting_session.public_token])
        ).json()
        self.assertEqual(state["queue"]["voted"], 1)
        self.assertEqual(state["queue"]["total"], 2)
        self.assertEqual(state["queue"]["missing"], 1)

    def test_organizer_sees_full_participant_progress(self):
        second_task = Task.objects.create(
            project=self.project, number="ABS-102", title="Вторая задача"
        )
        VotingSessionTask.objects.bulk_create(
            [
                VotingSessionTask(
                    session=self.voting_session, task=self.task, position=1
                ),
                VotingSessionTask(
                    session=self.voting_session, task=second_task, position=2
                ),
            ]
        )
        self.join_participant("Не начинал")
        in_progress = self.join_participant("В процессе")
        all_voted = self.join_participant("Оценил всё")
        completed = self.join_participant("Завершил")
        self.organizer.post(
            reverse("poker:session_start", args=[self.voting_session.pk])
        )
        vote_url = reverse(
            "poker:room_vote", args=[self.voting_session.public_token]
        )
        navigate_url = reverse(
            "poker:room_navigate", args=[self.voting_session.public_token]
        )
        complete_url = reverse(
            "poker:room_complete", args=[self.voting_session.public_token]
        )

        in_progress.post(vote_url, {"value": 4})
        for participant_client in (all_voted, completed):
            participant_client.post(vote_url, {"value": 8})
            participant_client.post(navigate_url, {"direction": "next"})
            participant_client.post(vote_url, {"value": 12})
        completed.post(complete_url)

        state = self.organizer.get(
            reverse("poker:session_state", args=[self.voting_session.pk])
        ).json()
        progress = {
            item["name"]: (item["progress_status"], item["progress_label"])
            for item in state["participants"]
        }
        self.assertEqual(
            progress["Не начинал"],
            ("not_started", "0 из 2 · не приступил"),
        )
        self.assertEqual(progress["В процессе"], ("in_progress", "1 из 2"))
        self.assertEqual(
            progress["Оценил всё"], ("all_voted", "2 из 2 · готов")
        )
        self.assertEqual(
            progress["Завершил"], ("completed", "2 из 2 · завершил")
        )
        self.assertEqual(state["completed_participant_count"], 1)
        self.assertEqual(state["participant_count"], 4)
        organizer_page = self.organizer.get(self.voting_session.get_absolute_url())
        self.assertContains(organizer_page, "0 из 2 · не приступил")
        self.assertContains(organizer_page, "1 из 2")
        self.assertContains(organizer_page, "2 из 2 · готов")
        self.assertContains(organizer_page, "2 из 2 · завершил")
        self.assertContains(organizer_page, "Завершили полностью: 1 из 4")
        self.assertContains(organizer_page, "Итог сохранён:")
        self.assertContains(organizer_page, "Сохранённая оценка")

    def test_personal_link_restores_existing_participant_on_another_device(self):
        second_task = Task.objects.create(
            project=self.project, number="ABS-102", title="Вторая задача"
        )
        VotingSessionTask.objects.bulk_create(
            [
                VotingSessionTask(
                    session=self.voting_session, task=self.task, position=1
                ),
                VotingSessionTask(
                    session=self.voting_session, task=second_task, position=2
                ),
            ]
        )
        original_client = self.join_participant("Анна")
        self.organizer.post(
            reverse("poker:session_start", args=[self.voting_session.pk])
        )
        original_client.post(
            reverse("poker:room_vote", args=[self.voting_session.public_token]),
            {"value": 8},
        )
        original_client.post(
            reverse(
                "poker:room_navigate",
                args=[self.voting_session.public_token],
            ),
            {"direction": "next"},
        )
        participant = Participant.objects.get(name="Анна")
        resume_url = reverse(
            "poker:room_resume",
            args=[self.voting_session.public_token, participant.client_token],
        )

        another_device = Client()
        response = another_device.get(resume_url)

        self.assertRedirects(
            response,
            reverse("poker:room", args=[self.voting_session.public_token]),
        )
        room_response = another_device.get(
            reverse("poker:room", args=[self.voting_session.public_token])
        )
        self.assertContains(room_response, "Вы вошли как <strong>Анна</strong>")
        self.assertContains(room_response, "Ссылка для продолжения")
        state = another_device.get(
            reverse("poker:room_state", args=[self.voting_session.public_token])
        ).json()
        self.assertEqual(state["current_task"]["id"], second_task.pk)
        self.assertEqual(state["queue"]["voted"], 1)
        self.assertEqual(self.voting_session.participants.count(), 1)
        self.assertEqual(Vote.objects.count(), 1)

    def test_invalid_personal_link_returns_to_join_without_new_participant(self):
        self.join_participant("Анна")
        invalid_url = reverse(
            "poker:room_resume",
            args=[self.voting_session.public_token, uuid.uuid4()],
        )

        response = Client().get(invalid_url, follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Персональная ссылка недействительна")
        self.assertContains(response, "Как вас представить?")
        self.assertEqual(self.voting_session.participants.count(), 1)

    def test_organizer_can_rotate_resume_link_without_losing_votes(self):
        participant_client = self.join_participant("Анна")
        self.organizer.post(
            reverse(
                "poker:session_start_task",
                args=[self.voting_session.pk, self.task.pk],
            )
        )
        participant_client.post(
            reverse("poker:room_vote", args=[self.voting_session.public_token]),
            {"value": 12},
        )
        participant = Participant.objects.get(name="Анна")
        old_token = participant.client_token
        old_resume_url = reverse(
            "poker:room_resume",
            args=[self.voting_session.public_token, old_token],
        )
        rotate_url = reverse(
            "poker:participant_resume_rotate",
            args=[self.voting_session.pk, participant.pk],
        )

        response = self.organizer.post(rotate_url)

        self.assertEqual(response.status_code, 200)
        participant.refresh_from_db()
        self.assertNotEqual(participant.client_token, old_token)
        self.assertEqual(Vote.objects.count(), 1)
        self.assertEqual(self.voting_session.participants.count(), 1)
        self.assertIn(str(participant.client_token), response.json()["resume_url"])
        self.assertEqual(
            participant_client.get(
                reverse(
                    "poker:room_state",
                    args=[self.voting_session.public_token],
                )
            ).status_code,
            403,
        )

        old_link_response = Client().get(old_resume_url, follow=True)
        self.assertContains(
            old_link_response, "Персональная ссылка недействительна"
        )
        new_device = Client()
        new_resume_url = reverse(
            "poker:room_resume",
            args=[
                self.voting_session.public_token,
                participant.client_token,
            ],
        )
        self.assertEqual(new_device.get(new_resume_url).status_code, 302)
        self.assertContains(
            new_device.get(
                reverse("poker:room", args=[self.voting_session.public_token])
            ),
            "Вы вошли как <strong>Анна</strong>",
        )

    def test_other_organizer_cannot_rotate_participant_resume_link(self):
        self.join_participant("Анна")
        participant = Participant.objects.get(name="Анна")
        original_token = participant.client_token
        other_user = get_user_model().objects.create_user(
            "other-owner", password="secret"
        )
        other_client = Client()
        other_client.force_login(other_user)

        response = other_client.post(
            reverse(
                "poker:participant_resume_rotate",
                args=[self.voting_session.pk, participant.pk],
            )
        )

        self.assertEqual(response.status_code, 404)
        participant.refresh_from_db()
        self.assertEqual(participant.client_token, original_token)

    def test_organizer_state_contains_participant_resume_controls(self):
        self.join_participant("Анна")
        participant = Participant.objects.get(name="Анна")

        state = self.organizer.get(
            reverse("poker:session_state", args=[self.voting_session.pk])
        ).json()

        participant_state = state["participants"][0]
        self.assertIn(str(participant.client_token), participant_state["resume_url"])
        self.assertEqual(
            participant_state["rotate_url"],
            reverse(
                "poker:participant_resume_rotate",
                args=[self.voting_session.pk, participant.pk],
            ),
        )

    def test_room_copy_keeps_queue_but_resets_voting_data(self):
        second_task = Task.objects.create(
            project=self.project, number="ABS-102", title="Вторая задача"
        )
        VotingSessionTask.objects.bulk_create(
            [
                VotingSessionTask(
                    session=self.voting_session, task=self.task, position=1
                ),
                VotingSessionTask(
                    session=self.voting_session, task=second_task, position=2
                ),
            ]
        )
        self.join_participant("Анна")

        response = self.organizer.post(
            reverse("poker:session_copy", args=[self.voting_session.pk])
        )

        copied = VotingSession.objects.exclude(pk=self.voting_session.pk).get()
        self.assertRedirects(response, copied.get_absolute_url())
        self.assertEqual(copied.status, VotingSession.Status.DRAFT)
        self.assertEqual(
            list(copied.queue_items.values_list("task_id", "position")),
            [(self.task.pk, 1), (second_task.pk, 2)],
        )
        self.assertEqual(copied.participants.count(), 0)
        self.assertEqual(copied.rounds.count(), 0)

    def test_participant_name_must_be_unique_in_room(self):
        self.join_participant("Анна")
        second = Client()
        response = second.post(
            reverse("poker:room_join", args=[self.voting_session.public_token]),
            {"name": "анна"},
        )
        self.assertEqual(response.status_code, 400)
        self.assertContains(
            response,
            "Участник с таким именем уже существует",
            status_code=400,
        )
        self.assertContains(
            response, "персональную ссылку продолжения", status_code=400
        )
        self.assertEqual(Participant.objects.count(), 1)


class SprintTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user("owner", password="secret")
        self.project = Project.objects.create(owner=self.user, name="ABS Core")
        self.task = Task.objects.create(
            project=self.project,
            number="ABS-10",
            title="Выгрузка для ЦБ",
            status=Task.Status.ESTIMATED,
            estimate_sum=14,
            estimate_count=3,
        )
        self.sprint = Sprint.objects.create(
            project=self.project, name="Спринт 24", capacity=20
        )
        SprintTask.objects.create(sprint=self.sprint, task=self.task, position=1)
        self.client.force_login(self.user)

    def test_sprint_total_uses_exact_task_average(self):
        self.assertEqual(self.sprint.total_estimate, self.task.estimate)
        self.assertEqual(self.sprint.total_estimate_display, "4.67")

    def test_sprint_statuses_and_archive_lifecycle(self):
        self.assertEqual(self.sprint.status, Sprint.Status.PLANNING)
        detail_response = self.client.get(self.sprint.get_absolute_url())
        self.assertContains(detail_response, "Планируется")
        self.assertContains(detail_response, "Активен")
        self.assertContains(detail_response, "Завершён")

        self.client.post(
            reverse("poker:sprint_set_status", args=[self.sprint.pk]),
            {"status": Sprint.Status.ACTIVE},
        )
        self.sprint.refresh_from_db()
        self.assertEqual(self.sprint.status, Sprint.Status.ACTIVE)

        self.client.post(reverse("poker:sprint_archive", args=[self.sprint.pk]))
        self.sprint.refresh_from_db()
        self.assertIsNone(self.sprint.archived_at)

        self.client.post(
            reverse("poker:sprint_set_status", args=[self.sprint.pk]),
            {"status": Sprint.Status.COMPLETED},
        )
        self.client.post(reverse("poker:sprint_archive", args=[self.sprint.pk]))
        self.sprint.refresh_from_db()
        self.assertEqual(self.sprint.status, Sprint.Status.COMPLETED)
        self.assertIsNotNone(self.sprint.archived_at)
        active_page = self.client.get(self.project.get_absolute_url())
        self.assertNotIn(self.sprint, list(active_page.context["sprints"]))
        archive_page = self.client.get(
            self.project.get_absolute_url(), {"archive": "1"}
        )
        self.assertIn(self.sprint, list(archive_page.context["sprints"]))

        self.client.post(reverse("poker:sprint_restore", args=[self.sprint.pk]))
        self.sprint.refresh_from_db()
        self.assertIsNone(self.sprint.archived_at)

    def test_completed_sprint_blocks_composition_changes_until_reopened(self):
        available_task = Task.objects.create(
            project=self.project,
            number="ABS-11",
            title="Дополнительная задача",
            status=Task.Status.ESTIMATED,
            estimate_sum=8,
            estimate_count=1,
        )
        self.sprint.status = Sprint.Status.COMPLETED
        self.sprint.save(update_fields=("status",))

        self.client.post(
            reverse("poker:sprint_add_tasks", args=[self.sprint.pk]),
            {"task_ids": [available_task.pk]},
        )
        self.assertFalse(
            SprintTask.objects.filter(
                sprint=self.sprint, task=available_task
            ).exists()
        )

        self.client.post(
            reverse("poker:sprint_remove_task", args=[self.sprint.pk, self.task.pk])
        )
        self.assertTrue(
            SprintTask.objects.filter(sprint=self.sprint, task=self.task).exists()
        )

        self.client.post(
            reverse("poker:sprint_set_status", args=[self.sprint.pk]),
            {"status": Sprint.Status.ACTIVE},
        )
        self.client.post(
            reverse("poker:sprint_add_tasks", args=[self.sprint.pk]),
            {"task_ids": [available_task.pk]},
        )
        self.assertTrue(
            SprintTask.objects.filter(
                sprint=self.sprint, task=available_task
            ).exists()
        )

    def test_organizer_can_delete_sprint_without_deleting_tasks(self):
        sprint_pk = self.sprint.pk
        task_pk = self.task.pk

        self.client.post(reverse("poker:sprint_delete", args=[sprint_pk]))

        self.assertFalse(Sprint.objects.filter(pk=sprint_pk).exists())
        self.assertTrue(Task.objects.filter(pk=task_pk).exists())

    def test_capacity_shows_remaining_and_overage(self):
        self.sprint.refresh_from_db()
        self.assertEqual(self.sprint.capacity_remaining_display, "15.33")
        self.assertFalse(self.sprint.is_over_capacity)

        large_task = Task.objects.create(
            project=self.project,
            number="ABS-11",
            title="Крупная задача",
            status=Task.Status.ESTIMATED,
            estimate_sum=20,
            estimate_count=1,
        )
        SprintTask.objects.create(
            sprint=self.sprint, task=large_task, position=2
        )
        sprint = Sprint.objects.get(pk=self.sprint.pk)

        self.assertTrue(sprint.is_over_capacity)
        self.assertEqual(sprint.capacity_overage_display, "4.67")
        self.assertEqual(sprint.capacity_remaining, Decimal("0"))
        response = self.client.get(sprint.get_absolute_url())
        self.assertContains(response, "Превышение: 4.67 points")
        self.assertContains(response, "План превышает заданную ёмкость")

    def test_competency_capacities_track_each_team_limit_separately(self):
        sprint = Sprint.objects.create(
            project=self.project,
            name="Компетенческий спринт",
            analysis_capacity=10,
            development_capacity=10,
            testing_capacity=6,
        )
        task_definitions = (
            ("A-1", Task.Competency.ANALYSIS, 8),
            ("D-1", Task.Competency.DEVELOPMENT, 12),
            ("T-1", Task.Competency.TESTING, 4),
            ("N-1", Task.Competency.NONE, 2),
        )
        for position, (number, competency, estimate) in enumerate(
            task_definitions, start=1
        ):
            task = Task.objects.create(
                project=self.project,
                number=number,
                title=f"Задача {number}",
                competency=competency,
                status=Task.Status.ESTIMATED,
                estimate_sum=estimate,
                estimate_count=1,
            )
            SprintTask.objects.create(
                sprint=sprint,
                task=task,
                position=position,
            )

        sprint = Sprint.objects.get(pk=sprint.pk)
        rows = {row["key"]: row for row in sprint.competency_capacity_rows}
        self.assertEqual(sprint.total_estimate, Decimal("26"))
        self.assertEqual(sprint.capacity_total, Decimal("26"))
        self.assertEqual(sprint.untyped_estimate, Decimal("2"))
        self.assertEqual(
            rows[Task.Competency.ANALYSIS]["remaining"], Decimal("2")
        )
        self.assertEqual(
            rows[Task.Competency.DEVELOPMENT]["overage"], Decimal("2")
        )
        self.assertEqual(
            rows[Task.Competency.TESTING]["remaining"], Decimal("2")
        )
        self.assertTrue(sprint.is_over_capacity)
        self.assertEqual(sprint.over_capacity_labels, ["Разработка"])

        response = self.client.get(sprint.get_absolute_url())
        self.assertContains(response, "Аналитика")
        self.assertContains(response, "8 из 10 points")
        self.assertContains(response, "12 из 10 points")
        self.assertContains(response, "Превышение: 2 points")
        self.assertContains(response, "Без типа: 2 points")

        export = self.client.get(reverse("poker:sprint_export", args=[sprint.pk]))
        workbook = load_workbook(BytesIO(export.content), data_only=False)
        capacity_sheet = workbook["Ёмкость"]
        self.assertEqual(capacity_sheet["A2"].value, "Аналитика")
        self.assertEqual(capacity_sheet["B2"].value, 8)
        self.assertEqual(capacity_sheet["C2"].value, 10)
        self.assertEqual(capacity_sheet["A5"].value, "Без типа")
        self.assertEqual(capacity_sheet["B5"].value, 2)

    def test_capacity_update_converts_legacy_total_and_validates_values(self):
        response = self.client.post(
            reverse("poker:sprint_capacity_update", args=[self.sprint.pk]),
            {
                "analysis_capacity": "8",
                "development_capacity": "20",
                "testing_capacity": "12",
            },
        )

        self.assertRedirects(response, self.sprint.get_absolute_url())
        self.sprint.refresh_from_db()
        self.assertIsNone(self.sprint.capacity)
        self.assertEqual(self.sprint.analysis_capacity, Decimal("8"))
        self.assertEqual(self.sprint.development_capacity, Decimal("20"))
        self.assertEqual(self.sprint.testing_capacity, Decimal("12"))

        invalid = self.client.post(
            reverse("poker:sprint_capacity_update", args=[self.sprint.pk]),
            {
                "analysis_capacity": "-1",
                "development_capacity": "20",
                "testing_capacity": "12",
            },
        )
        self.assertEqual(invalid.status_code, 400)
        self.sprint.refresh_from_db()
        self.assertEqual(self.sprint.analysis_capacity, Decimal("8"))

    def test_sprint_copy_keeps_planned_tasks_and_resets_dates(self):
        self.sprint.start_date = timezone.localdate()
        self.sprint.end_date = timezone.localdate() + timedelta(days=14)
        self.sprint.status = Sprint.Status.ACTIVE
        self.sprint.analysis_capacity = 8
        self.sprint.development_capacity = 20
        self.sprint.testing_capacity = 12
        self.sprint.save(
            update_fields=(
                "start_date",
                "end_date",
                "status",
                "analysis_capacity",
                "development_capacity",
                "testing_capacity",
            )
        )

        response = self.client.post(
            reverse("poker:sprint_copy", args=[self.sprint.pk])
        )

        copied = Sprint.objects.exclude(pk=self.sprint.pk).get()
        self.assertRedirects(response, copied.get_absolute_url())
        self.assertEqual(copied.status, Sprint.Status.PLANNING)
        self.assertIsNone(copied.start_date)
        self.assertIsNone(copied.end_date)
        self.assertEqual(copied.capacity, self.sprint.capacity)
        self.assertEqual(copied.analysis_capacity, self.sprint.analysis_capacity)
        self.assertEqual(
            copied.development_capacity, self.sprint.development_capacity
        )
        self.assertEqual(copied.testing_capacity, self.sprint.testing_capacity)
        self.assertEqual(
            list(
                copied.sprint_tasks.values_list("task_id", "status", "position")
            ),
            [(self.task.pk, SprintTask.Status.PLANNED, 1)],
        )

    def test_transfer_preserves_source_history_and_moves_capacity(self):
        target = Sprint.objects.create(
            project=self.project,
            name="Спринт 25",
            status=Sprint.Status.PLANNING,
            capacity=20,
        )

        response = self.client.post(
            reverse("poker:sprint_transfer_tasks", args=[self.sprint.pk]),
            {"target_sprint": target.pk, "task_ids": [self.task.pk]},
        )

        self.assertRedirects(response, self.sprint.get_absolute_url())
        source_item = SprintTask.objects.get(
            sprint=self.sprint, task=self.task
        )
        target_item = SprintTask.objects.get(sprint=target, task=self.task)
        self.assertEqual(source_item.status, SprintTask.Status.TRANSFERRED)
        self.assertEqual(source_item.transferred_to, target)
        self.assertIsNotNone(source_item.transferred_at)
        self.assertEqual(target_item.status, SprintTask.Status.PLANNED)

        source = Sprint.objects.get(pk=self.sprint.pk)
        target = Sprint.objects.get(pk=target.pk)
        self.assertEqual(source.total_estimate, Decimal("0"))
        self.assertEqual(target.total_estimate_display, "4.67")
        history_page = self.client.get(source.get_absolute_url())
        self.assertContains(history_page, "Перенесена")
        self.assertContains(history_page, target.name)

    def test_excel_export_contains_average_formula(self):
        self.task.competency = Task.Competency.ANALYSIS
        self.task.save(update_fields=("competency",))
        response = self.client.get(
            reverse("poker:sprint_export", args=[self.sprint.pk])
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        workbook = load_workbook(BytesIO(response.content), data_only=False)
        sheet = workbook["Задачи спринта"]
        self.assertEqual(sheet["B2"].value, "ABS-10")
        self.assertEqual(sheet["D2"].value, "=E2/F2")
        self.assertEqual(sheet["E2"].value, 14)
        self.assertEqual(sheet["F2"].value, 3)
        self.assertEqual(sheet["G2"].value, "Аналитика")
        capacity_sheet = workbook["Ёмкость"]
        self.assertEqual(capacity_sheet["A2"].value, "Общая ёмкость (старая версия)")
        self.assertEqual(capacity_sheet["C2"].value, 20)
